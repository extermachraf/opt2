"""
fix_recipe_macros.py
--------------------
For every recipe in Ricette per APP_v2.xlsx:
  - reads each ingredient's BDA-CODE + weight (grams)
  - looks up the per-100g macros in BDA_Banca Dati Alimentari_V1.2022.xlsx
  - calculates: macro_total = macro_per_100g * weight / 100
  - sums all ingredients → recipe totals
  - updates ONLY total_protein_g, total_carbs_g, total_fat_g, total_fiber_g
    in the Supabase `recipes` table  (total_calories is NOT touched)

Usage:
  python fix_recipe_macros.py            ← dry run (safe, no DB writes)
  python fix_recipe_macros.py --execute  ← actually write to DB
"""

import sys
import json
import unicodedata
import openpyxl
import requests

# ── Config ────────────────────────────────────────────────────────────────────
import os
_HERE        = os.path.dirname(os.path.abspath(__file__))
_ROOT        = os.path.join(_HERE, "..")          # project root
BDA_FILE     = os.path.join(_ROOT, "BDA_Banca Dati Alimentari_V1.2022.xlsx")
RICETTE_FILE = os.path.join(_ROOT, "Ricette per APP_v2.xlsx")
SUPABASE_URL = "https://jnukglbwreegcluvwgad.supabase.co"
SUPABASE_KEY = ("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
                ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImpudWtnbGJ3cmVlZ2NsdXZ3Z2FkIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzA5MTgyMTksImV4cCI6MjA4NjI3ODIxOX0"
                ".5LS97w7_UiiNj86D3yau9VFiy0vNWiKsG0wwb5OWb-I")
DRY_RUN = "--execute" not in sys.argv
# ─────────────────────────────────────────────────────────────────────────────


def safe_num(val):
    """BDA uses -2 / -3 for 'not available'. Treat those and None as 0."""
    if val is None:
        return 0.0
    try:
        f = float(val)
        return 0.0 if f < 0 else f
    except (TypeError, ValueError):
        return 0.0


def normalize(text):
    """Lowercase + remove accents for fuzzy title matching."""
    nfkd = unicodedata.normalize("NFKD", str(text))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


# ── 1. Load BDA ───────────────────────────────────────────────────────────────
def load_bda():
    print("Loading BDA …")
    wb = openpyxl.load_workbook(BDA_FILE, read_only=True, data_only=True)
    ws = wb["BDA-2022"]
    bda = {}
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx < 3:          # 3 header rows (Italian / English / units)
            continue
        code = row[1]        # col B = Codice Alimento
        if code is None:
            continue
        bda[str(code).strip()] = {
            "protein": safe_num(row[9]),    # Proteine totali
            "fat":     safe_num(row[12]),   # Lipidi totali
            "carbs":   safe_num(row[16]),   # Carboidrati disponibili (MSE)
            "fiber":   safe_num(row[19]),   # Fibra alimentare totale
        }
    wb.close()
    print(f"  {len(bda)} BDA entries loaded.")
    return bda


# ── 2. Parse Ricette Excel ────────────────────────────────────────────────────
def parse_sheet(ws):
    recipes, current = [], None
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:                      # header row
            continue
        name, weight, code = row[0], row[1], row[3]

        if name is None:
            continue
        name_s = str(name).strip()

        if name_s.lower().startswith("analisi per"):   # summary row → skip
            continue

        if code is not None:              # ingredient row
            if current is not None:
                try:
                    w = float(weight) if weight is not None else 0.0
                except (TypeError, ValueError):
                    w = 0.0
                current["ingredients"].append({
                    "code":   str(code).strip(),
                    "weight": w,
                    "name":   name_s,
                })
        else:                             # recipe title row (no BDA code)
            if weight is None or str(weight).strip() == "":
                if current:
                    recipes.append(current)
                current = {"title": name_s, "ingredients": []}

    if current:
        recipes.append(current)
    return recipes


def load_ricette():
    print("Loading Ricette …")
    wb   = openpyxl.load_workbook(RICETTE_FILE, read_only=True, data_only=True)
    all_ = []
    for sheet in ["Ricette PARTE1", "Ricette PARTE2"]:
        all_.extend(parse_sheet(wb[sheet]))
    wb.close()
    print(f"  {len(all_)} recipes parsed.")
    return all_


# ── 3. Calculate macros ───────────────────────────────────────────────────────
def resolve_code(code, bda):
    """Try exact code first; if not found and ends with _1, try _2 fallback."""
    if code in bda:
        return bda[code]
    if code.endswith("_1"):
        alt = code[:-2] + "_2"
        if alt in bda:
            return bda[alt]
    return None


def calc_macros(recipe, bda):
    p = c = f = fi = 0.0
    missing = []
    for ingr in recipe["ingredients"]:
        entry = resolve_code(ingr["code"], bda)
        if entry is None:
            missing.append(f'{ingr["name"]} ({ingr["code"]})')
            continue
        m = ingr["weight"] / 100.0
        p  += entry["protein"] * m
        c  += entry["carbs"]   * m
        f  += entry["fat"]     * m
        fi += entry["fiber"]   * m
    return {
        "total_protein_g": round(p,  2),
        "total_carbs_g":   round(c,  2),
        "total_fat_g":     round(f,  2),
        "total_fiber_g":   round(fi, 2),
        "missing":         missing,
    }


# ── 4. Fetch DB recipes ───────────────────────────────────────────────────────
def fetch_db_recipes():
    print("Fetching recipes from DB …")
    hdrs = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/recipes",
        params={"select": "id,title", "limit": "500"},
        headers=hdrs,
    )
    r.raise_for_status()
    db = r.json()
    print(f"  {len(db)} recipes in DB.")
    return {normalize(rec["title"]): rec["id"] for rec in db}


# ── 5. Update DB ──────────────────────────────────────────────────────────────
def update_recipe(recipe_id, macros):
    hdrs = {
        "apikey":       SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer":       "return=minimal",
    }
    payload = {k: macros[k] for k in
               ["total_protein_g", "total_carbs_g", "total_fat_g", "total_fiber_g"]}
    r = requests.patch(
        f"{SUPABASE_URL}/rest/v1/recipes",
        params={"id": f"eq.{recipe_id}"},
        headers=hdrs,
        data=json.dumps(payload),
    )
    return r.status_code in (200, 204), r.status_code


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    mode = "DRY RUN" if DRY_RUN else "EXECUTE"
    print(f"\n{'='*60}")
    print(f"  fix_recipe_macros.py  [{mode}]")
    print(f"{'='*60}\n")

    bda          = load_bda()
    ricette      = load_ricette()
    db_map       = fetch_db_recipes()   # {normalized_title: uuid}

    matched = unmatched_titles = updated = errors = 0
    all_missing_codes = {}

    for recipe in ricette:
        macros  = calc_macros(recipe, bda)
        norm_t  = normalize(recipe["title"])
        db_id   = db_map.get(norm_t)

        if db_id is None:
            print(f"[NO MATCH] '{recipe['title']}'")
            unmatched_titles += 1
            continue

        matched += 1
        print(f"\n[{'DRY' if DRY_RUN else 'UPD'}] {recipe['title']}")
        print(f"       protein={macros['total_protein_g']}g  "
              f"carbs={macros['total_carbs_g']}g  "
              f"fat={macros['total_fat_g']}g  "
              f"fiber={macros['total_fiber_g']}g")

        if macros["missing"]:
            print(f"       ⚠ BDA code not found: {macros['missing']}")
            all_missing_codes[recipe["title"]] = macros["missing"]

        if not DRY_RUN:
            ok, status = update_recipe(db_id, macros)
            if ok:
                updated += 1
                print(f"       ✓ updated (HTTP {status})")
            else:
                errors += 1
                print(f"       ✗ FAILED (HTTP {status})")

    print(f"\n{'='*60}")
    print(f"  Recipes matched  : {matched}")
    print(f"  Unmatched titles : {unmatched_titles}")
    if not DRY_RUN:
        print(f"  DB rows updated  : {updated}")
        print(f"  Errors           : {errors}")
    if all_missing_codes:
        print(f"\n  Ingredients with missing BDA codes:")
        for title, codes in all_missing_codes.items():
            print(f"    {title}: {codes}")
    print(f"{'='*60}\n")
    if DRY_RUN:
        print("  *** This was a DRY RUN — nothing was written to the DB. ***")
        print("  *** Run with --execute to apply changes.                ***\n")


if __name__ == "__main__":
    main()

